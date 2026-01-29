from datetime import datetime
from typing import List, Dict

from openpyxl import load_workbook


def import_xlsx(path) -> List[Dict]:
    workbook = load_workbook(path)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        level, sku, rule_type, text, created_at = row
        if not level or not rule_type or not text:
            continue
        created_date = None
        if created_at:
            created_date = datetime.strptime(str(created_at), "%Y-%m-%d").date()
        rows.append(
            {
                "level": level,
                "sku": sku,
                "type": rule_type,
                "text": text,
                "created_at": created_date,
            }
        )
    return rows
