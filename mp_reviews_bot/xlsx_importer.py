from dataclasses import dataclass, field
from datetime import datetime, timedelta
from io import BytesIO
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

DEFAULT_SHEET_NAME = "Sheet"
DEFAULT_XLSX_HEADERS = (
    "marketplace",
    "seller_sku",
    "marketplace_item_id",
    "internal_sku",
    "product_name",
)


@dataclass
class RebindResult:
    marketplace: str
    seller_sku: str
    old_internal_sku: str
    new_internal_sku: str


@dataclass
class InvalidRow:
    row_number: int
    reason: str


@dataclass
class ImportResult:
    valid_rows: List[Dict]
    invalid_rows: List[InvalidRow]
    rebinds: List[RebindResult]


class UploadRateLimiter:
    def __init__(self, window: timedelta, max_uploads: int) -> None:
        self.window = window
        self.max_uploads = max_uploads
        self._seen: Dict[str, List[datetime]] = {}

    def check(self, key: str) -> None:
        now = datetime.utcnow()
        timestamps = [t for t in self._seen.get(key, []) if now - t < self.window]
        if len(timestamps) >= self.max_uploads:
            raise ValueError("rate limit exceeded")
        timestamps.append(now)
        self._seen[key] = timestamps


class SkuXlsxImporter:
    def __init__(
        self,
        max_file_size_bytes: int,
        timeout_seconds: int,
        rate_limiter: Optional[UploadRateLimiter] = None,
    ) -> None:
        self.max_file_size_bytes = max_file_size_bytes
        self.timeout_seconds = timeout_seconds
        self.rate_limiter = rate_limiter

    def import_xlsx(
        self,
        file_bytes: bytes,
        existing_internal_by_seller: Optional[Dict[Tuple[str, str], str]] = None,
        rate_limit_key: Optional[str] = None,
    ) -> ImportResult:
        if len(file_bytes) > self.max_file_size_bytes:
            raise ValueError("file exceeds size limit")
        if self.rate_limiter and rate_limit_key:
            self.rate_limiter.check(rate_limit_key)

        workbook = load_workbook(filename=BytesIO(file_bytes))
        if DEFAULT_SHEET_NAME not in workbook.sheetnames:
            raise ValueError("missing sheet")
        sheet = workbook[DEFAULT_SHEET_NAME]
        header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if len(header_row) > len(DEFAULT_XLSX_HEADERS):
            raise ValueError("unexpected extra columns")
        if tuple(header_row) != DEFAULT_XLSX_HEADERS:
            raise ValueError("XLSX headers must match template")

        valid_rows: List[Dict] = []
        invalid_rows: List[InvalidRow] = []
        rebinds: List[RebindResult] = []
        existing_internal_by_seller = existing_internal_by_seller or {}

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            marketplace, seller_sku, marketplace_item_id, internal_sku, product_name = row
            if not marketplace or not seller_sku or not marketplace_item_id or not internal_sku:
                invalid_rows.append(InvalidRow(row_number=idx, reason="missing required"))
                continue
            key = (str(marketplace), str(seller_sku))
            old_internal = existing_internal_by_seller.get(key)
            if old_internal and old_internal != str(internal_sku):
                rebinds.append(
                    RebindResult(
                        marketplace=str(marketplace),
                        seller_sku=str(seller_sku),
                        old_internal_sku=old_internal,
                        new_internal_sku=str(internal_sku),
                    )
                )
            valid_rows.append(
                {
                    "marketplace": str(marketplace),
                    "seller_sku": str(seller_sku),
                    "marketplace_item_id": str(marketplace_item_id),
                    "internal_sku": str(internal_sku),
                    "product_name": str(product_name) if product_name else None,
                }
            )
        return ImportResult(valid_rows=valid_rows, invalid_rows=invalid_rows, rebinds=rebinds)
