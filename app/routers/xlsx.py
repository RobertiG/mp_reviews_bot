from io import BytesIO
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app import crud
from app.db import get_db

router = APIRouter(prefix="/xlsx", tags=["xlsx"])

MAX_FILE_SIZE = 5 * 1024 * 1024


@router.post("/import/{project_id}")
async def import_xlsx(project_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large")
    wb = load_workbook(filename=BytesIO(content))
    sheet = wb.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    required = ["marketplace", "seller_sku", "marketplace_item_id", "internal_sku", "product_name"]
    if headers[:5] != required:
        raise HTTPException(status_code=400, detail="Invalid XLSX template")
    imported = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        marketplace, seller_sku, marketplace_item_id, internal_sku, product_name = row
        if not (marketplace and seller_sku and marketplace_item_id and internal_sku):
            continue
        crud.upsert_sku_map(
            db,
            project_id,
            str(marketplace),
            str(seller_sku),
            str(marketplace_item_id),
            str(internal_sku),
            str(product_name) if product_name else None,
        )
        imported += 1
    return {"imported": imported}
